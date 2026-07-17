from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import DetailSnapshotSummary


GROUPS = (
    ("denominator", "统计范围", "本次指标计算纳入的全部记录"),
    ("numerator", "达到要求", "在本院口径规定时间或条件内达到要求的记录"),
    ("unmatched", "未达到要求", "已纳入统计范围、但未达到本院口径要求的记录"),
)


def safe_excel_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    if isinstance(value, int) and abs(value) >= 10**15:
        return str(value)
    return value


def _group_rows(rows: list[dict[str, Any]], group: str) -> list[dict[str, Any]]:
    if group == "denominator":
        return rows
    expected = 1 if group == "numerator" else 0
    return [
        row
        for row in rows
        if int(row.get("__meets_numerator") or 0) == expected
    ]


def _version_text(summary: DetailSnapshotSummary) -> str:
    if summary.effective_level == "hospital" and summary.hospital_version is not None:
        standard = f"；标准版本 v{summary.national_version}" if summary.national_version else ""
        return f"本院口径 v{summary.hospital_version}{standard}"
    return f"标准口径 v{summary.national_version or '-'}"


def _field_lineage_text(summary: DetailSnapshotSummary) -> str:
    lines: list[str] = []
    for item in summary.field_lineage:
        if item.kind == "column" and item.sources:
            source = item.sources[0]
            lines.append(f"{item.label} → {source}")
        else:
            lines.append(f"{item.label} → {item.explanation}")
    return "\n".join(lines) or "未记录"


def _write_sheet(
    workbook: Workbook,
    *,
    title: str,
    description: str,
    summary: DetailSnapshotSummary,
    rows: list[dict[str, Any]],
    actor_id: str,
    first: bool,
) -> None:
    sheet = workbook.active if first else workbook.create_sheet()
    sheet.title = f"{title}_{len(rows)}"
    metadata = (
        ("指标名称", summary.rule_name),
        ("指标编号", summary.rule_id),
        ("适用医院", summary.hospital_id),
        ("口径来源与版本", _version_text(summary)),
        ("来源数据库", summary.source_database or "未记录"),
        ("取数表", "、".join(summary.source_tables) or "未记录"),
        ("字段来源", _field_lineage_text(summary)),
        ("统计区间", f"{summary.stat_start} 至 {summary.stat_end}（不含结束时刻）"),
        ("明细快照时间", summary.created_at.isoformat(sep=" ", timespec="seconds")),
        ("导出人", actor_id),
        ("本表说明", description),
        ("记录总数", len(rows)),
    )
    for index, (label, value) in enumerate(metadata, start=1):
        sheet.cell(index, 1, label)
        sheet.cell(index, 2, safe_excel_value(value))
        sheet.cell(index, 1).font = Font(bold=True)
        if label == "字段来源":
            sheet.cell(index, 2).alignment = Alignment(
                vertical="top", wrap_text=True
            )
            sheet.row_dimensions[index].height = max(
                30, 15 * max(1, len(summary.field_lineage))
            )

    labels = [column.label for column in summary.columns] + ["是否达到要求"]
    header_row = len(metadata) + 2
    data_start_row = header_row + 1
    for column_index, label in enumerate(labels, start=1):
        cell = sheet.cell(header_row, column_index, label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="087F78")
        cell.alignment = Alignment(horizontal="center")
    for row_index, row in enumerate(rows, start=data_start_row):
        values = [row.get(column.field) for column in summary.columns]
        values.append("是" if int(row.get("__meets_numerator") or 0) == 1 else "否")
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_index, column_index, safe_excel_value(value))
    sheet.freeze_panes = f"A{data_start_row}"
    last_row = max(header_row, data_start_row + len(rows) - 1)
    sheet.auto_filter.ref = (
        f"A{header_row}:{sheet.cell(last_row, len(labels)).coordinate}"
    )
    for index, label in enumerate(labels, start=1):
        sheet.column_dimensions[sheet.cell(header_row, index).column_letter].width = min(
            36, max(14, len(label) * 2 + 4)
        )
    sheet.column_dimensions["B"].width = max(
        float(sheet.column_dimensions["B"].width or 0), 36
    )


def create_indicator_workbook(
    path: Path,
    summary: DetailSnapshotSummary,
    rows: list[dict[str, Any]],
    *,
    actor_id: str,
) -> Path:
    numerator_count = sum(
        1 for row in rows if int(row.get("__meets_numerator") or 0) == 1
    )
    if (
        len(rows) != summary.denominator_count
        or numerator_count != summary.numerator_count
        or len(rows) - numerator_count != summary.unmatched_count
    ):
        raise ValueError("快照明细与聚合数量不一致，不能生成 Excel")
    workbook = Workbook()
    for index, (group, title, description) in enumerate(GROUPS):
        _write_sheet(
            workbook,
            title=title,
            description=description,
            summary=summary,
            rows=_group_rows(rows, group),
            actor_id=actor_id,
            first=index == 0,
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


_COMPARISON_HEADERS = (
    "对比项",
    "系统值",
    "上传文件值",
    "差异（上传文件值 - 系统值）",
    "单位",
    "结论",
    "上传文件来源列",
)


def _write_comparison_table(
    sheet,
    metrics: list[dict[str, Any]],
    *,
    header_row: int,
) -> None:
    for column_index, label in enumerate(_COMPARISON_HEADERS, start=1):
        cell = sheet.cell(header_row, column_index, label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="087F78")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_index, item in enumerate(metrics, start=header_row + 1):
        values = (
            item.get("metric"),
            item.get("system_value"),
            item.get("uploaded_value"),
            item.get("difference"),
            item.get("unit"),
            "一致" if item.get("match") else "不一致",
            item.get("source_column"),
        )
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_index, column_index, safe_excel_value(value))
    sheet.freeze_panes = f"A{header_row + 1}"
    last_row = max(header_row, header_row + len(metrics))
    sheet.auto_filter.ref = f"A{header_row}:G{last_row}"
    widths = (14, 14, 16, 30, 12, 12, 22)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(header_row, index).column_letter].width = width


def create_upload_comparison_workbook(
    path: Path,
    comparison: dict[str, Any],
    *,
    actor_id: str,
    created_at: datetime,
) -> Path:
    """生成上传文件与系统试运行结果的汇总级或逐条差异工作簿。"""
    if comparison.get("comparison_level") == "row":
        return _create_row_comparison_workbook(
            path,
            comparison,
            actor_id=actor_id,
            created_at=created_at,
        )
    metrics = list(comparison.get("metrics") or [])
    if not metrics:
        raise ValueError("未识别到可对比的分子、分母或指标率")
    workbook = Workbook()
    summary = workbook.active
    summary.title = "对比摘要"
    metadata = (
        ("指标名称", comparison.get("rule_name") or comparison.get("rule_id") or "未记录"),
        ("指标编号", comparison.get("rule_id") or "未记录"),
        ("适用医院", comparison.get("hospital_id") or "未记录"),
        ("系统统计区间", comparison.get("system_stat_period") or "未记录"),
        ("上传文件", comparison.get("file_name") or "未记录"),
        ("对比层级", "汇总级"),
        ("差异方向", comparison.get("comparison_direction") or "上传文件值 - 系统值"),
        ("逐条对比说明", comparison.get("row_level_note") or "未记录"),
        ("一致项", int(comparison.get("matched_count") or 0)),
        ("不一致项", int(comparison.get("different_count") or 0)),
        ("导出人", actor_id),
        ("导出时间", created_at.isoformat(sep=" ", timespec="seconds")),
    )
    for row_index, (label, value) in enumerate(metadata, start=1):
        summary.cell(row_index, 1, label).font = Font(bold=True)
        summary.cell(row_index, 2, safe_excel_value(value))
        summary.cell(row_index, 2).alignment = Alignment(vertical="top", wrap_text=True)
    _write_comparison_table(summary, metrics, header_row=len(metadata) + 2)
    summary.column_dimensions["B"].width = max(
        summary.column_dimensions["B"].width or 0, 42
    )

    groups = (
        ("一致项", [item for item in metrics if item.get("match")]),
        ("不一致项", [item for item in metrics if not item.get("match")]),
    )
    for title, rows in groups:
        sheet = workbook.create_sheet(f"{title}_{len(rows)}")
        sheet.cell(1, 1, "说明").font = Font(bold=True)
        sheet.cell(
            1,
            2,
            "这里只比较上传文件与系统试运行的汇总值，不代表患者级记录交集或差集。",
        )
        sheet.cell(1, 2).alignment = Alignment(wrap_text=True)
        _write_comparison_table(sheet, rows, header_row=3)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _write_record_table(
    sheet,
    headers: list[str],
    rows: list[list[Any]],
    *,
    description: str,
) -> None:
    sheet.cell(1, 1, "说明").font = Font(bold=True)
    sheet.cell(1, 2, description)
    sheet.cell(1, 2).alignment = Alignment(wrap_text=True)
    header_row = 3
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column_index, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="087F78")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_index, values in enumerate(rows, start=header_row + 1):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_index, column_index, safe_excel_value(value))
    sheet.freeze_panes = "A4"
    last_row = max(header_row, header_row + len(rows))
    last_column = max(1, len(headers))
    sheet.auto_filter.ref = (
        f"A{header_row}:{sheet.cell(last_row, last_column).coordinate}"
    )
    for index, header in enumerate(headers, start=1):
        width = 22 if header not in {"匹配键", "字段差异"} else 36
        sheet.column_dimensions[sheet.cell(header_row, index).column_letter].width = width


def _create_row_comparison_workbook(
    path: Path,
    comparison: dict[str, Any],
    *,
    actor_id: str,
    created_at: datetime,
) -> Path:
    matched = list(comparison.get("matched_rows") or [])
    system_only = list(comparison.get("system_only_rows") or [])
    uploaded_only = list(comparison.get("uploaded_only_rows") or [])
    common_fields = list(comparison.get("common_fields") or [])
    system_fields = list(dict.fromkeys(
        common_fields
        + list(comparison.get("system_only_fields") or [])
        + [key for row in system_only for key in row]
    ))
    uploaded_fields = list(dict.fromkeys(
        common_fields
        + list(comparison.get("uploaded_only_fields") or [])
        + [key for row in uploaded_only for key in row]
    ))

    workbook = Workbook()
    summary = workbook.active
    summary.title = "对比摘要"
    metadata = (
        ("指标名称", comparison.get("rule_name") or comparison.get("rule_id") or "未记录"),
        ("指标编号", comparison.get("rule_id") or "未记录"),
        ("适用医院", comparison.get("hospital_id") or "未记录"),
        ("系统统计区间", comparison.get("system_stat_period") or "未记录"),
        ("上传文件统计区间", comparison.get("uploaded_stat_period") or "未记录"),
        ("上传文件", comparison.get("file_name") or "未记录"),
        ("对比层级", "逐条记录"),
        ("逐条匹配字段", "、".join(comparison.get("matching_fields") or []) or "未记录"),
        ("双方都有", len(matched)),
        ("仅系统有", len(system_only)),
        ("仅上传文件有", len(uploaded_only)),
        ("同一记录但字段值不同", int(comparison.get("field_difference_count") or 0)),
        ("系统达到要求记录", int(comparison.get("system_numerator_count") or 0)),
        ("上传文件达到要求记录", int(comparison.get("uploaded_numerator_count") or 0)),
        ("同一记录但达标判定不同", int(comparison.get("classification_difference_count") or 0)),
        ("已确认差异", "\n".join(comparison.get("confirmed_findings") or []) or "无"),
        ("导出人", actor_id),
        ("导出时间", created_at.isoformat(sep=" ", timespec="seconds")),
    )
    for row_index, (label, value) in enumerate(metadata, start=1):
        summary.cell(row_index, 1, label).font = Font(bold=True)
        summary.cell(row_index, 2, safe_excel_value(value))
        summary.cell(row_index, 2).alignment = Alignment(vertical="top", wrap_text=True)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 72

    matched_headers = (
        ["匹配键", "字段差异"]
        + [f"系统-{field}" for field in system_fields]
        + [f"上传文件-{field}" for field in uploaded_fields]
    )
    matched_values = [
        [item.get("key"), "、".join(item.get("different_fields") or []) or "无"]
        + [item.get("system", {}).get(field) for field in system_fields]
        + [item.get("uploaded", {}).get(field) for field in uploaded_fields]
        for item in matched
    ]
    matched_sheet = workbook.create_sheet(f"双方都有_{len(matched)}")
    _write_record_table(
        matched_sheet,
        matched_headers,
        matched_values,
        description="按匹配字段识别为同一业务记录；字段差异列会列出值不一致的字段。",
    )

    system_sheet = workbook.create_sheet(f"仅系统有_{len(system_only)}")
    _write_record_table(
        system_sheet,
        system_fields,
        [[row.get(field) for field in system_fields] for row in system_only],
        description="当前系统试运行明细中存在、上传文件中未匹配到的记录。",
    )

    uploaded_sheet = workbook.create_sheet(f"仅上传文件有_{len(uploaded_only)}")
    _write_record_table(
        uploaded_sheet,
        uploaded_fields,
        [[row.get(field) for field in uploaded_fields] for row in uploaded_only],
        description="上传文件中存在、当前系统试运行明细中未匹配到的记录。",
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path
