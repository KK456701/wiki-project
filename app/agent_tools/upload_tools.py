"""上传 Excel 指标明细的模型可见分析工具。"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timezone
from functools import partial
from pathlib import Path
import re
from typing import Any, Callable

from pydantic import BaseModel, ConfigDict, Field

from app.agent_runtime.contracts import AgentRunState, AgentRuntimeContext
from app.agent_tools.contracts import AgentTool, ToolEvidence, ToolResult, ToolRiskLevel

_UPLOAD_ROOT = Path(__file__).resolve().parents[2] / "runtime" / "uploads"
_UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)


class AnalyzeUploadedIndicatorsInput(BaseModel):
    model_config = ConfigDict(extra="forbid")

    file_key: str = Field(min_length=1, max_length=128)


@dataclass(frozen=True, slots=True)
class UploadToolServices:
    detail_loader: Callable[[str, str, str], dict[str, Any]] | None = None


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _detail_header_index(rows: list[tuple[Any, ...]]) -> int | None:
    for index, row in enumerate(rows):
        headers = {_cell_text(value) for value in row if _cell_text(value)}
        if "是否达到要求" in headers and len(headers) >= 2:
            return index
    return None


def _sheet_metadata(
    rows: list[tuple[Any, ...]], header_index: int | None
) -> dict[str, Any]:
    if header_index is None:
        return {}
    metadata: dict[str, Any] = {}
    for row in rows[:header_index]:
        if len(row) < 2:
            continue
        label = _cell_text(row[0])
        if label:
            metadata[label] = row[1]
    return metadata


def _rule_id_from_file(file_path: Path, metadata: dict[str, Any]) -> str:
    configured = _cell_text(metadata.get("指标编号") or metadata.get("指标编码"))
    if configured:
        return configured
    match = re.search(r"(?:^|_)([A-Za-z]+\d{4}_\d+)(?:_|\.)", file_path.name)
    return match.group(1).upper() if match else ""


def parse_excel_preview(file_path: Path) -> dict[str, Any]:
    """解析 Excel 并返回摘要数据，不暴露患者明细。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl 未安装，无法解析 Excel。"}

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        return {"error": f"Excel 文件无法打开：{exc}"}

    sheets_info: list[dict[str, Any]] = []
    total_rows = 0
    headers_sample: dict[str, list[str]] = {}
    detail_datasets: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 5001), values_only=True))
        if not rows:
            continue

        header_index = _detail_header_index(rows)
        actual_header_index = header_index if header_index is not None else 0
        headers = [_cell_text(cell) for cell in rows[actual_header_index]]
        data_rows = [
            row
            for row in rows[actual_header_index + 1 :]
            if any(value is not None and _cell_text(value) for value in row)
        ]
        sheet_total = len(data_rows)
        total_rows += sheet_total

        # 数值列统计
        numeric_stats: dict[str, dict[str, float]] = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            values = [
                float(row[col_idx])
                for row in data_rows
                if col_idx < len(row) and row[col_idx] is not None
                and isinstance(row[col_idx], (int, float))
            ]
            if values:
                numeric_stats[header] = {
                    "min": round(min(values), 4),
                    "max": round(max(values), 4),
                    "sum": round(sum(values), 4),
                    "avg": round(sum(values) / len(values), 4),
                    "count": len(values),
                }

        sheets_info.append({
            "sheet_name": sheet_name,
            "headers": headers[:30],
            "row_count": sheet_total,
            "numeric_columns": numeric_stats,
            "metadata": _sheet_metadata(rows, header_index),
        })
        if header_index is not None:
            detail_datasets.append({
                "sheet_name": sheet_name,
                "headers": headers,
                "rows": [
                    {
                        header: row[column_index] if column_index < len(row) else None
                        for column_index, header in enumerate(headers)
                        if header
                    }
                    for row in data_rows
                ],
                "metadata": _sheet_metadata(rows, header_index),
            })
        if headers and sheet_name not in headers_sample:
            headers_sample[sheet_name] = headers[:30]

    # 尝试检测指标结构
    indicator_hints = _detect_indicator_structure(headers_sample, sheets_info)

    primary_detail = next(
        (
            dataset
            for dataset in detail_datasets
            if str(dataset["sheet_name"]).startswith("统计范围")
        ),
        detail_datasets[0] if detail_datasets else None,
    )
    export_metadata: dict[str, Any] | None = None
    if primary_detail is not None:
        metadata = dict(primary_detail.get("metadata") or {})
        export_metadata = {
            "rule_id": _rule_id_from_file(file_path, metadata),
            "rule_name": _cell_text(metadata.get("指标名称")),
            "hospital_id": _cell_text(metadata.get("适用医院")),
            "stat_period": _cell_text(metadata.get("统计区间")),
            "record_count": len(primary_detail.get("rows") or []),
        }
        total_rows = export_metadata["record_count"]

    result = {
        "file_name": file_path.name,
        "sheet_count": len(sheets_info),
        "total_rows": total_rows,
        "sheets": sheets_info,
        "indicator_hints": indicator_hints,
        "detail_export": export_metadata,
        "_detail_dataset": primary_detail,
    }
    wb.close()
    return result


def public_excel_preview(preview: dict[str, Any]) -> dict[str, Any]:
    """移除患者行，仅保留可发送给模型的工作簿结构和汇总信息。"""
    return {
        key: value
        for key, value in preview.items()
        if not key.startswith("_")
    }


def _detect_indicator_structure(
    headers_sample: dict[str, list[str]],
    sheets_info: list[dict[str, Any]],
) -> dict[str, Any]:
    """检测 Excel 中可能是指标相关结构的列。"""
    all_headers = set()
    for headers in headers_sample.values():
        all_headers.update(h.lower() for h in headers if h)

    hints: dict[str, Any] = {}

    # 常见指标列名
    numerator_keywords = {"分子", "numerator", "num"}
    denominator_keywords = {"分母", "denominator", "denom"}
    rate_keywords = {"指标率", "率", "rate", "ratio", "percentage", "比例"}
    period_keywords = {"统计周期", "时间", "period", "date", "月份", "季度", "年份"}
    name_keywords = {"指标名称", "indicator", "name", "名称"}

    found = {
        "numerator_cols": [h for h in all_headers if any(kw in h for kw in numerator_keywords)],
        "denominator_cols": [h for h in all_headers if any(kw in h for kw in denominator_keywords)],
        "rate_cols": [h for h in all_headers if any(kw in h for kw in rate_keywords)],
        "period_cols": [h for h in all_headers if any(kw in h for kw in period_keywords)],
        "name_cols": [h for h in all_headers if any(kw in h for kw in name_keywords)],
    }
    hints["detected_columns"] = found
    hints["looks_like_indicator_data"] = (
        bool(found["numerator_cols"] or found["rate_cols"])
        and bool(found["denominator_cols"] or found["rate_cols"])
    )

    # 尝试提取指标率数值
    for sheet in sheets_info:
        numeric = sheet.get("numeric_columns") or {}
        potential_rates = []
        for col_name, stats in numeric.items():
            col_lower = col_name.lower()
            if any(kw in col_lower for kw in rate_keywords):
                potential_rates.append({"column": col_name, "stats": stats, "role": "rate"})
            elif any(kw in col_lower for kw in numerator_keywords):
                potential_rates.append({"column": col_name, "stats": stats, "role": "numerator"})
            elif any(kw in col_lower for kw in denominator_keywords):
                potential_rates.append({"column": col_name, "stats": stats, "role": "denominator"})
        if potential_rates:
            hints.setdefault("potential_indicator_values", []).extend(potential_rates)

    return hints


def _stat_value(stats: dict[str, Any]) -> float | None:
    value = stats.get("avg")
    if value is None:
        value = stats.get("sum")
    return float(value) if value is not None else None


def build_aggregate_comparison(
    preview: dict[str, Any],
    system_result: dict[str, Any],
) -> dict[str, Any]:
    """构造汇总级对比，不把单行汇总误报成患者级差异。"""
    hints = preview.get("indicator_hints") or {}
    candidates = hints.get("potential_indicator_values") or []
    values_by_role: dict[str, float] = {}
    columns_by_role: dict[str, str] = {}
    for candidate in candidates:
        role = str(candidate.get("role") or "")
        if role not in {"numerator", "denominator", "rate"} or role in values_by_role:
            continue
        value = _stat_value(candidate.get("stats") or {})
        if value is None:
            continue
        values_by_role[role] = value
        columns_by_role[role] = str(candidate.get("column") or "")

    definitions = (
        ("denominator", "分母", "system_denominator", "人次", 0.01),
        ("numerator", "分子", "system_numerator", "人次", 0.01),
        ("rate", "指标率", "system_rate", "百分点", 0.01),
    )
    metrics: list[dict[str, Any]] = []
    for role, label, system_key, unit, tolerance in definitions:
        if role not in values_by_role or system_result.get(system_key) is None:
            continue
        uploaded = values_by_role[role]
        system = float(system_result[system_key])
        difference = round(uploaded - system, 4)
        metrics.append({
            "metric": label,
            "role": role,
            "source_column": columns_by_role[role],
            "system_value": round(system, 4),
            "uploaded_value": round(uploaded, 4),
            "difference": difference,
            "unit": unit,
            "match": abs(difference) < tolerance,
        })

    return {
        "comparison_level": "aggregate",
        "comparison_direction": "上传文件值 - 系统值",
        "row_level_comparison_available": False,
        "cause_analysis_available": False,
        "confirmed_causes": [],
        "row_level_note": (
            "上传文件仅包含汇总值，未提供入院流水号等逐条标识，"
            "因此只能核对分子、分母和指标率，不能判断具体记录的交集与差集。"
        ),
        "cause_analysis_note": (
            "当前文件没有逐条业务记录，无法确认差异是否由重复记录、统计周期、"
            "ICU 排除、时间计算或字段映射造成；这些因素不能作为本次对比结论。"
        ),
        "required_fields_for_cause_analysis": [
            "admission_id",
            "admit_time",
            "transfer_time",
            "from_dept_id",
            "to_dept_id",
        ],
        "system_stat_period": system_result.get("system_stat_period"),
        "metrics": metrics,
        "matched_count": sum(1 for item in metrics if item["match"]),
        "different_count": sum(1 for item in metrics if not item["match"]),
    }


def _canonical_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="milliseconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}[ T].*", text):
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
            return parsed.isoformat(timespec="milliseconds")
        except ValueError:
            pass
    return text


def _comparison_key_fields(common_headers: list[str]) -> list[str]:
    identity_terms = (
        "患者标识", "入院流水号", "admission_id", "申请编号", "记录编号",
        "consult_id", "transfer_id", "request_id",
    )
    event_time_terms = (
        "申请时间", "入院时间", "转科时间", "发生时间", "request_time",
        "admit_time", "transfer_time",
    )
    identity = [
        header for header in common_headers
        if any(term.lower() in header.lower() for term in identity_terms)
    ]
    event_times = [
        header for header in common_headers
        if any(term.lower() in header.lower() for term in event_time_terms)
    ]
    selected = list(dict.fromkeys(identity + event_times))
    if selected:
        return selected
    return [
        header for header in common_headers
        if header not in {"是否达到要求", "到位耗时（分钟）", "转科耗时（分钟）"}
    ]


def _row_key(row: dict[str, Any], fields: list[str]) -> tuple[str, ...]:
    return tuple(_canonical_value(row.get(field)) for field in fields)


def _meets_requirement(row: dict[str, Any]) -> bool:
    return _canonical_value(row.get("是否达到要求")).lower() in {
        "是", "1", "true", "yes", "y",
    }


def _system_label_rows(system_details: dict[str, Any]) -> list[dict[str, Any]]:
    columns = list(system_details.get("columns") or [])
    rows: list[dict[str, Any]] = []
    for raw in system_details.get("rows") or []:
        item = {
            str(column.get("label") or column.get("field")): raw.get(column.get("field"))
            for column in columns
            if column.get("field")
        }
        item["是否达到要求"] = (
            "是" if int(raw.get("__meets_numerator") or 0) == 1 else "否"
        )
        rows.append(item)
    return rows


def build_row_level_comparison(
    preview: dict[str, Any],
    system_details: dict[str, Any],
    *,
    include_rows: bool = False,
) -> dict[str, Any]:
    """按稳定业务键执行多重集合对比，重复记录不会被 set 静默去重。"""
    uploaded = preview.get("detail_export") or {}
    uploaded_rule_id = _cell_text(uploaded.get("rule_id"))
    uploaded_rule_name = _cell_text(uploaded.get("rule_name"))
    system_rule_id = _cell_text(system_details.get("rule_id"))
    system_rule_name = _cell_text(system_details.get("rule_name"))
    base = {
        "system_rule_id": system_rule_id,
        "system_rule_name": system_rule_name,
        "uploaded_rule_id": uploaded_rule_id,
        "uploaded_rule_name": uploaded_rule_name,
        "system_stat_period": _cell_text(system_details.get("stat_period")),
        "uploaded_stat_period": _cell_text(uploaded.get("stat_period")),
    }
    if not uploaded_rule_id or not system_rule_id:
        return {
            **base,
            "comparison_status": "identity_missing",
            "row_level_comparison_available": False,
            "message": "上传文件或系统结果缺少指标编号，不能进行逐条对比。",
        }
    if uploaded_rule_id != system_rule_id:
        return {
            **base,
            "comparison_status": "indicator_mismatch",
            "row_level_comparison_available": False,
            "message": (
                f"上传文件属于“{uploaded_rule_name or uploaded_rule_id}”({uploaded_rule_id})，"
                f"当前查询属于“{system_rule_name or system_rule_id}”({system_rule_id})，"
                "两个指标不能进行汇总或逐条差异比较。"
            ),
        }

    dataset = preview.get("_detail_dataset") or {}
    uploaded_rows = list(dataset.get("rows") or [])
    system_rows = _system_label_rows(system_details)
    uploaded_headers = [str(value) for value in dataset.get("headers") or []]
    system_headers = [
        str(column.get("label") or column.get("field"))
        for column in system_details.get("columns") or []
    ] + ["是否达到要求"]
    common_headers = [header for header in system_headers if header in uploaded_headers]
    key_fields = _comparison_key_fields(common_headers)
    if not key_fields:
        return {
            **base,
            "comparison_status": "matching_fields_missing",
            "row_level_comparison_available": False,
            "message": "两个文件没有可用于识别同一业务记录的公共字段。",
        }

    uploaded_by_key: dict[tuple[str, ...], list[dict[str, Any]]] = {}
    for row in uploaded_rows:
        uploaded_by_key.setdefault(_row_key(row, key_fields), []).append(row)

    matched: list[dict[str, Any]] = []
    system_only: list[dict[str, Any]] = []
    for system_row in system_rows:
        key = _row_key(system_row, key_fields)
        candidates = uploaded_by_key.get(key) or []
        if not candidates:
            system_only.append(system_row)
            continue
        uploaded_row = candidates.pop(0)
        differences = [
            header
            for header in common_headers
            if _canonical_value(system_row.get(header))
            != _canonical_value(uploaded_row.get(header))
        ]
        matched.append({
            "key": " | ".join(key),
            "system": system_row,
            "uploaded": uploaded_row,
            "different_fields": differences,
        })

    uploaded_only = [
        row
        for candidates in uploaded_by_key.values()
        for row in candidates
    ]
    system_period = _cell_text(system_details.get("stat_period"))
    uploaded_period = _cell_text(uploaded.get("stat_period"))
    confirmed_findings: list[str] = []
    if system_period and uploaded_period and system_period != uploaded_period:
        confirmed_findings.append(
            f"统计区间不一致：系统为 {system_period}；上传文件为 {uploaded_period}。"
        )
    if system_only:
        confirmed_findings.append(f"有 {len(system_only)} 条记录仅存在于系统结果。")
    if uploaded_only:
        confirmed_findings.append(f"有 {len(uploaded_only)} 条记录仅存在于上传文件。")
    changed = sum(1 for item in matched if item["different_fields"])
    system_numerator_count = sum(1 for row in system_rows if _meets_requirement(row))
    uploaded_numerator_count = sum(1 for row in uploaded_rows if _meets_requirement(row))
    system_only_numerator_count = sum(
        1 for row in system_only if _meets_requirement(row)
    )
    uploaded_only_numerator_count = sum(
        1 for row in uploaded_only if _meets_requirement(row)
    )
    classification_difference_count = sum(
        1
        for item in matched
        if _meets_requirement(item["system"]) != _meets_requirement(item["uploaded"])
    )
    if changed:
        confirmed_findings.append(f"双方匹配记录中有 {changed} 条存在字段值差异。")
    confirmed_findings.append(
        "分母记录差异已拆分为："
        f"仅系统有 {len(system_only)} 条、仅上传文件有 {len(uploaded_only)} 条。"
    )
    confirmed_findings.append(
        f"达到要求记录：系统 {system_numerator_count} 条、上传文件 {uploaded_numerator_count} 条；"
        f"其中仅系统有 {system_only_numerator_count} 条、仅上传文件有 "
        f"{uploaded_only_numerator_count} 条，双方同一记录但判定不同 "
        f"{classification_difference_count} 条。"
    )

    result = {
        **base,
        "comparison_status": "row_level_compared",
        "comparison_level": "row",
        "row_level_comparison_available": True,
        "matching_fields": key_fields,
        "common_fields": common_headers,
        "system_only_fields": [header for header in system_headers if header not in uploaded_headers],
        "uploaded_only_fields": [header for header in uploaded_headers if header not in system_headers],
        "system_count": len(system_rows),
        "uploaded_count": len(uploaded_rows),
        "both_count": len(matched),
        "system_only_count": len(system_only),
        "uploaded_only_count": len(uploaded_only),
        "field_difference_count": changed,
        "system_numerator_count": system_numerator_count,
        "uploaded_numerator_count": uploaded_numerator_count,
        "system_only_numerator_count": system_only_numerator_count,
        "uploaded_only_numerator_count": uploaded_only_numerator_count,
        "classification_difference_count": classification_difference_count,
        "confirmed_findings": confirmed_findings,
    }
    if include_rows:
        result.update({
            "matched_rows": matched,
            "system_only_rows": system_only,
            "uploaded_only_rows": uploaded_only,
        })
    return result


def _save_upload(file_content: bytes, filename: str, hospital_id: str) -> str:
    safe_name = f"{hospital_id}_{_utcnow().strftime('%Y%m%d%H%M%S')}_{filename}"
    file_path = _UPLOAD_ROOT / safe_name
    file_path.write_bytes(file_content)
    return safe_name


def analyze_uploaded_indicators(
    arguments: AnalyzeUploadedIndicatorsInput,
    context: AgentRuntimeContext,
    state: AgentRunState,
    services: UploadToolServices,
) -> ToolResult:
    file_path = _UPLOAD_ROOT / arguments.file_key
    if not file_path.exists() or not file_path.is_file():
        return ToolResult(
            ok=False,
            status="not_found",
            code="UPLOAD_NOT_FOUND",
            summary="未找到已上传的文件，请先上传 Excel 文件。",
        )

    # 安全检查：只能访问自己医院的文件
    if not file_path.name.startswith(f"{context.hospital_id}_"):
        return ToolResult(
            ok=False,
            status="forbidden",
            code="UPLOAD_ACCESS_DENIED",
            summary="无权访问其他医院的上传文件。",
        )

    preview = parse_excel_preview(file_path)
    if "error" in preview:
        return ToolResult(
            ok=False,
            status="error",
            code="EXCEL_PARSE_ERROR",
            summary=preview["error"],
        )

    # 构造与当前试运行结果的对比
    comparison = {}
    trial_run_id = ""
    current_rule_id = ""
    for result in reversed(state.last_tool_results):
        if not isinstance(result, dict):
            continue
        data = result.get("data") or {}
        if isinstance(data, dict) and data.get("stat_start"):
            trial_run_id = _cell_text(data.get("run_id"))
            current_rule_id = _cell_text(data.get("rule_id"))
            comparison = {
                "system_stat_period": f"{data.get('stat_start')} 至 {data.get('stat_end')}",
                "system_numerator": data.get("numerator_count"),
                "system_denominator": data.get("denominator_count"),
                "system_rate": data.get("result_value"),
            }
            break

    hints = preview.get("indicator_hints") or {}
    current_rule_name = ""
    for result in reversed(state.last_tool_results):
        data = result.get("data") if isinstance(result, dict) else None
        if not isinstance(data, dict):
            continue
        if _cell_text(data.get("rule_id")) == current_rule_id:
            current_rule_name = _cell_text(data.get("rule_name")) or current_rule_name

    aggregate_comparison = None
    row_comparison = None
    if comparison and preview.get("detail_export"):
        system_details = {
            "rule_id": current_rule_id,
            "rule_name": current_rule_name or current_rule_id,
            "stat_period": comparison.get("system_stat_period"),
            "columns": [],
            "rows": [],
        }
        uploaded_rule_id = _cell_text((preview.get("detail_export") or {}).get("rule_id"))
        if (
            uploaded_rule_id == current_rule_id
            and services.detail_loader is not None
            and "indicator_detail_view" in context.permissions
            and trial_run_id
        ):
            system_details = services.detail_loader(
                trial_run_id, context.hospital_id, context.user_id
            )
        row_comparison = build_row_level_comparison(preview, system_details)
    elif comparison:
        aggregate_comparison = build_aggregate_comparison(preview, comparison)
    comparisons_detail = (
        [
            {
                "column": item["source_column"],
                "role": item["role"],
                "excel_value": item["uploaded_value"],
                "system_value": item["system_value"],
                "difference": item["difference"],
                "match": item["match"],
            }
            for item in aggregate_comparison["metrics"]
        ]
        if aggregate_comparison
        else []
    )

    public_preview = public_excel_preview(preview)
    analysis_summary = f"已解析 {preview['file_name']}，共 {preview['total_rows']} 行数据。"
    if row_comparison and row_comparison.get("comparison_status") == "indicator_mismatch":
        analysis_summary = str(row_comparison.get("message") or analysis_summary)

    return ToolResult(
        ok=True,
        status="success",
        code="UPLOAD_ANALYZED",
        summary=analysis_summary,
        data={
            "file_name": preview["file_name"],
            "file_key": arguments.file_key,
            "sheet_count": preview["sheet_count"],
            "total_rows": preview["total_rows"],
            "sheets": [
                {
                    "name": s["sheet_name"],
                    "headers": s["headers"],
                    "row_count": s["row_count"],
                }
                for s in public_preview["sheets"]
            ],
            "detail_export": public_preview.get("detail_export"),
            "indicator_hints": hints,
            "comparison_with_system": comparison if comparison else None,
            "aggregate_comparison": aggregate_comparison,
            "row_comparison": row_comparison,
            "comparisons_detail": comparisons_detail if comparisons_detail else None,
        },
        evidence=[ToolEvidence(
            source="uploaded_excel",
            source_id=arguments.file_key,
            fact_types=["file_analysis"],
        )],
    )


def build_upload_tools(services: UploadToolServices) -> list[AgentTool]:
    return [
        AgentTool(
            name="analyze_uploaded_indicators",
            description=(
                "分析已上传的医院指标 Excel 文件，提取表头、行数、数值列统计（最小/最大/均值/总和），"
                "自动检测指标相关列（分子/分母/指标率/统计周期），先校验上传文件与当前结果的指标身份；"
                "同一指标且上传的是系统明细导出时，按稳定业务键统计双方都有、仅系统有、仅上传文件有及字段差异，"
                "只向模型输出脱敏汇总证据，不暴露患者明细。需要先通过上传接口提交文件获得 file_key。"
            ),
            input_model=AnalyzeUploadedIndicatorsInput,
            handler=partial(analyze_uploaded_indicators, services=services),
            risk_level=ToolRiskLevel.READ,
            timeout_seconds=30.0,
        ),
    ]
