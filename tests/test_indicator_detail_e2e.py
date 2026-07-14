from __future__ import annotations

from io import BytesIO
from pathlib import Path

import yaml
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.pool import StaticPool

from app.api import hospital_auth as auth_api
from app.api import indicator_details as detail_api
from app import config as app_config
from app.db_access.query_result import QueryResult
from app.hospital_auth import dependencies as auth_dependencies
from app.hospital_auth.repository import HospitalAuthRepository
from app.hospital_auth.schema import ensure_hospital_auth_schema
from app.hospital_auth.service import HospitalAuthService
from app.indicator_details.repository import IndicatorDetailRepository
from app.indicator_details.schema import ensure_indicator_detail_schema
from app.indicator_details.service import IndicatorDetailService
from app.indicator_details.snapshot import DetailSnapshotStore
from tests.test_indicator_detail_sql import make_context


ROOT = Path(__file__).resolve().parents[1]


class _FakeBusinessDB:
    source_id = "hospital_demo_data"
    tool_name = "execute_sql_hospital_demo_data"

    def __init__(self) -> None:
        self.rows = [
            {
                "patient_id": f"PATIENT{index:04d}",
                "dept_id": "ED" if index % 2 else "ICU",
                "consult_type": "急会诊",
                "request_time": f"2026-07-{(index % 28) + 1:02d} 09:00:00",
                "arrive_time": f"2026-07-{(index % 28) + 1:02d} 09:15:00",
                "arrive_minutes": 15 if index <= 488 else 31,
                "__meets_numerator": 1 if index <= 488 else 0,
                "__evidence_row_count": 1,
            }
            for index in range(1, 577)
        ]

    def execute_select(self, sql: str) -> QueryResult:
        assert sql.lstrip().upper().startswith("SELECT")
        return QueryResult(
            rows=[dict(row) for row in self.rows],
            row_count=len(self.rows),
            source=self.source_id,
            tool_name=self.tool_name,
            duration_ms=5,
        )


def _make_client(tmp_path: Path) -> TestClient:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE med_sql_run_log (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  run_id TEXT NOT NULL UNIQUE,
                  sql_id TEXT,
                  hospital_id TEXT NOT NULL,
                  rule_id TEXT NOT NULL,
                  stat_start_time TEXT,
                  stat_end_time TEXT,
                  run_status TEXT NOT NULL,
                  result_value REAL,
                  error_message TEXT,
                  duration_ms INTEGER,
                  run_by TEXT,
                  numerator_count INTEGER,
                  denominator_count INTEGER,
                  run_context_json TEXT,
                  run_time TEXT NOT NULL
                )
                """
            )
        )
    ensure_hospital_auth_schema(engine)
    ensure_indicator_detail_schema(engine)

    context = make_context("MQSI2025_005")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO med_sql_run_log
                  (run_id, sql_id, hospital_id, rule_id, stat_start_time,
                   stat_end_time, run_status, result_value, numerator_count,
                   denominator_count, run_context_json, run_time)
                VALUES
                  ('RUN_E2E_001', 'SQL_E2E_001', 'hospital_001',
                   'MQSI2025_005', :stat_start, :stat_end, 'success', 84.72,
                   488, 576, :run_context, '2026-07-14 09:00:00')
                """
            ),
            {
                "stat_start": context.stat_start,
                "stat_end": context.stat_end,
                "run_context": context.model_dump_json(),
            },
        )

    auth_repository = HospitalAuthRepository(engine)
    auth_service = HospitalAuthService(auth_repository)
    auth_service.create_or_reset_local_user(
        account_id="user_001",
        hospital_id="hospital_001",
        password="123456",
        permissions={"indicator_detail_view", "indicator_detail_export"},
        must_change_password=True,
    )
    export_root = tmp_path / "exports"
    detail_repository = IndicatorDetailRepository(engine)
    detail_service = IndicatorDetailService(
        detail_repository,
        DetailSnapshotStore(
            detail_repository,
            _FakeBusinessDB(),
            export_root=export_root,
        ),
        auth_repository,
        export_root=export_root,
    )

    app = FastAPI()
    app.include_router(auth_api.router)
    app.include_router(detail_api.router)
    app.dependency_overrides[
        auth_dependencies.get_hospital_auth_service
    ] = lambda: auth_service
    app.dependency_overrides[
        detail_api.get_indicator_detail_service
    ] = lambda: detail_service
    return TestClient(app)


def _login_and_change_password(client: TestClient) -> dict[str, str]:
    login = client.post(
        "/api/auth/hospital/login",
        json={"account_id": "user_001", "password": "123456"},
    )
    assert login.status_code == 200
    assert login.json()["must_change_password"] is True
    changed = client.post(
        "/api/auth/hospital/change-password",
        headers={"Authorization": f"Bearer {login.json()['token']}"},
        json={"current_password": "123456", "new_password": "Hospital2026"},
    )
    assert changed.status_code == 200
    assert changed.json()["must_change_password"] is False
    return {"Authorization": f"Bearer {changed.json()['token']}"}


def test_delivery_config_and_readme_cover_detail_acceptance() -> None:
    assert app_config._CONFIG_PATH == ROOT / "config.yaml"
    config = yaml.safe_load(
        (ROOT / "config.example.yaml").read_text(encoding="utf-8")
    )
    assert config["hospital_auth_session_hours"] == 8
    assert config["indicator_detail_export_root"] == "runtime/exports"
    assert config["indicator_detail_expire_hours"] == 24
    assert config["indicator_detail_max_rows"] == 20_000
    assert config["indicator_detail_default_page_size"] == 50

    readme = (ROOT / "README.md").read_text(encoding="utf-8")
    assert "指标明细预览与短期导出验收" in readme
    assert "indicator_detail_view" in readme
    assert "页面预览脱敏，Excel 保留授权完整值" in readme


def test_urgent_consult_preview_and_excel_are_consistent(tmp_path: Path) -> None:
    client = _make_client(tmp_path)
    headers = _login_and_change_password(client)

    snapshot = client.post("/api/sql-runs/RUN_E2E_001/details", headers=headers)
    assert snapshot.status_code == 201
    assert {
        "denominator": snapshot.json()["denominator_count"],
        "numerator": snapshot.json()["numerator_count"],
        "unmatched": snapshot.json()["unmatched_count"],
    } == {"denominator": 576, "numerator": 488, "unmatched": 88}

    preview = client.get(
        "/api/sql-runs/RUN_E2E_001/details/numerator?page=1&page_size=50",
        headers=headers,
    )
    assert preview.status_code == 200
    assert preview.json()["total"] == 488
    patient_id = preview.json()["items"][0]["患者标识"]
    assert "*" in patient_id
    assert "PATIENT0001" not in preview.text

    created = client.post(
        "/api/sql-runs/RUN_E2E_001/exports",
        headers=headers,
        json={"confirmed": True},
    )
    assert created.status_code == 201
    downloaded = client.get(
        f"/api/indicator-exports/{created.json()['export_id']}/download",
        headers=headers,
    )
    assert downloaded.status_code == 200
    workbook = load_workbook(BytesIO(downloaded.content), read_only=True)
    assert workbook.sheetnames == [
        "统计范围_576",
        "达到要求_488",
        "未达到要求_88",
    ]
    assert workbook["统计范围_576"]["B1"].value == "急会诊及时到位率"
    workbook.close()
