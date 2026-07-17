from __future__ import annotations

import base64
from datetime import datetime, timedelta
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, text

from app.hospital_auth.models import HospitalPrincipal
from app.indicator_details.models import DetailColumn, DetailSnapshotSummary, RunContext
from app.indicator_details.repository import IndicatorDetailRepository
from app.indicator_details.schema import ensure_indicator_detail_schema
from app.indicator_details.service import IndicatorDetailError, IndicatorDetailService


class _Clock:
    def __init__(self) -> None:
        self.now = datetime(2026, 7, 14, 9, 0, 0)

    def __call__(self) -> datetime:
        return self.now


class _AuditRepository:
    def __init__(self) -> None:
        self.items: list[dict] = []

    def insert_audit(self, **item) -> None:
        self.items.append(item)


class _SnapshotStore:
    def __init__(self, summary: DetailSnapshotSummary) -> None:
        self.summary = summary
        self.rows = [
            {"patient_id": "PATIENT001", "__meets_numerator": 1},
            {"patient_id": "PATIENT002", "__meets_numerator": 1},
            {"patient_id": "PATIENT003", "__meets_numerator": 0},
        ]

    def create(self, run_id, hospital_id, actor_id):
        return self.summary

    def read_all_rows(self, run_id, hospital_id):
        return self.summary, self.rows


def _principal(hospital_id: str = "hospital_001", *, export: bool = True):
    permissions = {"indicator_detail_view"}
    if export:
        permissions.add("indicator_detail_export")
    return HospitalPrincipal(
        user_id="USER_001",
        account_id="user_001",
        hospital_id=hospital_id,
        permissions=frozenset(permissions),
        must_change_password=False,
        session_id="SESSION_001",
    )


def _service(tmp_path: Path, *, export_ttl: timedelta = timedelta(hours=24)):
    engine = create_engine(f"sqlite:///{tmp_path / 'runtime.db'}")
    with engine.begin() as conn:
        conn.execute(
            text(
                "CREATE TABLE med_sql_run_log ("
                "id INTEGER PRIMARY KEY AUTOINCREMENT, run_id VARCHAR(64), "
                "hospital_id VARCHAR(64), rule_id VARCHAR(64), run_status VARCHAR(32), "
                "stat_start_time VARCHAR(32), stat_end_time VARCHAR(32), "
                "result_value FLOAT, numerator_count INTEGER, denominator_count INTEGER, "
                "run_context_json TEXT)"
            )
        )
        conn.execute(
            text(
                "INSERT INTO med_sql_run_log (run_id, hospital_id, rule_id) "
                "VALUES ('RUN_001', 'hospital_001', 'MQSI2025_005')"
            )
        )
        conn.execute(text(
            "UPDATE med_sql_run_log SET run_status='success', "
            "stat_start_time='2026-01-01 00:00:00', "
            "stat_end_time='2026-07-17 00:00:00', result_value=2.83, "
            "numerator_count=11, denominator_count=389, "
            "run_context_json='{\"effective_rule\":{\"rule_name\":\"急会诊及时到位率\"}}'"
        ))
    ensure_indicator_detail_schema(engine)
    repository = IndicatorDetailRepository(engine)
    clock = _Clock()
    summary = DetailSnapshotSummary(
        snapshot_id="SNAP_001",
        run_id="RUN_001",
        hospital_id="hospital_001",
        rule_id="MQSI2025_005",
        rule_name="急会诊及时到位率",
        effective_level="hospital",
        national_version="2025",
        hospital_version=1,
        stat_start="2026-07-01 00:00:00",
        stat_end="2026-08-01 00:00:00",
        denominator_count=3,
        numerator_count=2,
        unmatched_count=1,
        columns=[
            DetailColumn(field="patient_id", label="患者标识", sensitivity="patient_id")
        ],
        created_at=clock.now,
        expires_at=clock.now + timedelta(hours=24),
    )
    audit = _AuditRepository()
    service = IndicatorDetailService(
        repository,
        _SnapshotStore(summary),
        audit,
        export_root=tmp_path / "exports",
        upload_root=tmp_path / "uploads",
        now_provider=clock,
        export_ttl=export_ttl,
    )
    return service, repository, audit, clock


def test_service_hides_other_hospital_run_and_requires_export_permission(tmp_path: Path) -> None:
    service, _, audit, _ = _service(tmp_path)

    with pytest.raises(IndicatorDetailError) as cross_hospital:
        service.ensure_snapshot(_principal("hospital_002"), "RUN_001")
    with pytest.raises(IndicatorDetailError) as no_permission:
        service.create_export(_principal(export=False), "RUN_001", True)

    assert cross_hospital.value.status_code == 404
    assert no_permission.value.status_code == 403
    assert [item["action"] for item in audit.items] == ["ACCESS_DENIED", "ACCESS_DENIED"]


def test_service_hides_context_validation_internals(tmp_path: Path) -> None:
    service, _, audit, _ = _service(tmp_path)

    def invalid_context(*_args):
        return RunContext.model_validate({"unexpected": "internal value"})

    service.snapshot_store.create = invalid_context

    with pytest.raises(IndicatorDetailError) as error:
        service.ensure_snapshot(_principal(), "RUN_001")

    assert error.value.code == "DETAIL_CONTEXT_INVALID"
    assert "validation errors" not in str(error.value)
    assert "重新试运行" in str(error.value)
    assert audit.items[-1]["reason"] == "DETAIL_CONTEXT_INVALID"


def test_service_creates_downloads_and_expires_excel(tmp_path: Path) -> None:
    service, repository, audit, clock = _service(tmp_path)
    principal = _principal()

    with pytest.raises(IndicatorDetailError) as confirmation:
        service.create_export(principal, "RUN_001", False)
    assert confirmation.value.status_code == 400

    export = service.create_export(principal, "RUN_001", True)
    record = repository.get_export(export.export_id)
    path = service._resolve_relative_path(record["relative_path"])
    workbook = load_workbook(path, read_only=True)

    assert workbook.sheetnames == ["统计范围_3", "达到要求_2", "未达到要求_1"]
    workbook.close()
    downloaded_path, _ = service.resolve_download(principal, export.export_id)
    assert downloaded_path == path
    assert repository.get_export(export.export_id)["download_count"] == 1

    clock.now += timedelta(hours=25)
    cleanup = service.cleanup_expired()

    assert cleanup.expired_exports == 1
    assert not path.exists()
    assert repository.get_export(export.export_id)["status"] == "expired"
    assert {item["action"] for item in audit.items} >= {
        "DETAIL_EXPORT_CREATE",
        "DETAIL_EXPORT_DOWNLOAD",
        "DETAIL_FILE_EXPIRED",
    }


def test_export_expiry_can_be_configured(tmp_path: Path) -> None:
    service, _, _, clock = _service(tmp_path, export_ttl=timedelta(hours=2))

    export = service.create_export(_principal(), "RUN_001", True)

    assert export.expires_at == clock.now + timedelta(hours=2)


def test_service_creates_upload_comparison_export(tmp_path: Path) -> None:
    service, repository, audit, _ = _service(tmp_path)
    upload_root = tmp_path / "uploads"
    upload_root.mkdir()
    file_key = "hospital_001_report.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["denominator", "numerator", "rate_pct"])
    sheet.append([522, 30, 5.75])
    workbook.save(upload_root / file_key)
    token = base64.urlsafe_b64encode(file_key.encode("utf-8")).decode("ascii").rstrip("=")

    export = service.create_upload_comparison_export(
        _principal(), "RUN_001", token, True
    )
    record = repository.get_export(export.export_id)
    path = service._resolve_relative_path(record["relative_path"])
    result = load_workbook(path, read_only=False, data_only=False)

    assert result.sheetnames == ["对比摘要", "一致项_0", "不一致项_3"]
    assert result["对比摘要"]["D15"].value == 133
    assert result["不一致项_3"]["A4"].value == "分母"
    assert record["row_count"] == 3
    assert audit.items[-1]["action"] == "UPLOAD_COMPARISON_EXPORT_CREATE"


def _write_detail_upload(path: Path, rule_id: str, rule_name: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "统计范围_2"
    for row in (
        ["指标名称", rule_name],
        ["指标编号", rule_id],
        ["适用医院", "hospital_001"],
        ["统计区间", "2026-07-01 00:00:00 至 2026-08-01 00:00:00"],
        [],
        ["患者标识", "是否达到要求"],
        ["PATIENT001", "是"],
        ["PATIENT004", "否"],
    ):
        sheet.append(row)
    workbook.save(path)


def test_service_rejects_detail_export_from_another_indicator(tmp_path: Path) -> None:
    service, _, _, _ = _service(tmp_path)
    upload_root = tmp_path / "uploads"
    upload_root.mkdir()
    file_key = "hospital_001_MQSI2025_001_detail.xlsx"
    _write_detail_upload(upload_root / file_key, "MQSI2025_001", "患者入院48小时内转科的比例")
    token = base64.urlsafe_b64encode(file_key.encode("utf-8")).decode("ascii").rstrip("=")

    with pytest.raises(IndicatorDetailError) as error:
        service.create_upload_comparison_export(_principal(), "RUN_001", token, True)

    assert error.value.code == "UPLOAD_COMPARISON_INDICATOR_MISMATCH"
    assert "两个指标不能" in str(error.value)


def test_service_creates_row_level_comparison_export_for_same_indicator(tmp_path: Path) -> None:
    service, repository, _, _ = _service(tmp_path)
    upload_root = tmp_path / "uploads"
    upload_root.mkdir()
    file_key = "hospital_001_MQSI2025_005_detail.xlsx"
    _write_detail_upload(upload_root / file_key, "MQSI2025_005", "急会诊及时到位率")
    token = base64.urlsafe_b64encode(file_key.encode("utf-8")).decode("ascii").rstrip("=")

    export = service.create_upload_comparison_export(
        _principal(), "RUN_001", token, True
    )
    record = repository.get_export(export.export_id)
    path = service._resolve_relative_path(record["relative_path"])
    workbook = load_workbook(path, read_only=False, data_only=False)

    assert workbook.sheetnames == [
        "对比摘要",
        "双方都有_1",
        "仅系统有_2",
        "仅上传文件有_1",
    ]
    assert record["row_count"] == 4
